"""Convert a SportsbookReviewsOnline (SBRO) NBA odds export into the JSON feed
that ``ingest_odds`` consumes (``make ingest-odds FILE=...``).

SBRO is the deepest *free* closing-line archive (~2007→present); see
``docs/RESOURCES.md``. Its bulk Excel (export to CSV) lays each game across **two
consecutive rows** — the visitor first, then the home team — with columns like::

    Date, Rot, VH, Team, 1st, 2nd, 3rd, 4th, Final, Open, Close, ML, 2H

We read the ``Date``, ``VH`` (V/H), ``Team``, and ``ML`` (American moneyline)
columns, pair each V row with the H row that follows it, and emit::

    {"sport": "nba", "date": "YYYY-MM-DD",
     "home_team": "...", "away_team": "...",
     "home_close": <ML>, "away_close": <ML>}

The moneyline is passed through verbatim (``ingest_odds._to_decimal`` handles
American→decimal and the vig guard), so this converter stays a pure reshaper:
no odds math, no I/O in the core. SBRO uses terse one-token city labels
(``GoldenState``, ``NewYork``); :data:`SBRO_TEAMS` maps them to full
``"City Mascot"`` names so the downstream ``canonical_event_id`` collapses onto
the same row as the nba_api-ingested events. Unmapped labels pass through raw
(``normalize_team`` still reduces them) and are counted in the run log.

SBRO has no year in the date cell (a season spans two calendar years), so pass
``--season-start-year`` (the year the season *tipped off*, e.g. 2023 for
2023-24). Months >= July map to that year; Jan–Jun to the next.

A second, friendlier source is the **SBRO mirror** (``flancast90/
sportsbookreview-scraper``), whose ``data/nba_archive_10Y.json`` is already one
pre-joined record per game with a full ``YYYYMMDD`` date and American close
moneylines. Pass it with ``--format archive`` (auto-detected from a ``.json``
extension); no ``--season-start-year`` needed. See :func:`archive_json_to_records`.
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from ..logging_conf import get_logger

log = get_logger("sbro_to_feed")

# SBRO's terse team label -> full "City Mascot" name. Downstream normalize_team
# reduces these to the canonical mascot token (Trail Blazers / 76ers / LA* are
# handled by matching.TWO_WORD_MASCOTS + ALIASES), so the converted feed keys
# onto the same canonical_event_id as the free nba_api history. Includes a few
# historical labels for the deep seasons SBRO covers.
SBRO_TEAMS = {
    "Atlanta": "Atlanta Hawks", "Boston": "Boston Celtics",
    "Brooklyn": "Brooklyn Nets", "Charlotte": "Charlotte Hornets",
    "Chicago": "Chicago Bulls", "Cleveland": "Cleveland Cavaliers",
    "Dallas": "Dallas Mavericks", "Denver": "Denver Nuggets",
    "Detroit": "Detroit Pistons", "GoldenState": "Golden State Warriors",
    "Houston": "Houston Rockets", "Indiana": "Indiana Pacers",
    "LAClippers": "LA Clippers", "LALakers": "Los Angeles Lakers",
    "Memphis": "Memphis Grizzlies", "Miami": "Miami Heat",
    "Milwaukee": "Milwaukee Bucks", "Minnesota": "Minnesota Timberwolves",
    "NewOrleans": "New Orleans Pelicans", "NewYork": "New York Knicks",
    "OklahomaCity": "Oklahoma City Thunder", "Orlando": "Orlando Magic",
    "Philadelphia": "Philadelphia 76ers", "Phoenix": "Phoenix Suns",
    "Portland": "Portland Trail Blazers", "Sacramento": "Sacramento Kings",
    "SanAntonio": "San Antonio Spurs", "Toronto": "Toronto Raptors",
    "Utah": "Utah Jazz", "Washington": "Washington Wizards",
    # historical labels (deep SBRO seasons)
    "Seattle": "Seattle SuperSonics", "NewJersey": "New Jersey Nets",
    "NewOrleansOklahoma": "New Orleans Pelicans", "VancouverGrizzlies": "Memphis Grizzlies",
    "GoldenStateWarriors": "Golden State Warriors",
}

# The flancast/FinnedAI SBRO *mirror* (`*_archive_10Y.json`) labels teams by bare
# mascot, which normalize_team already resolves — except a handful of multi-word
# city tokens that would reduce to the wrong word ("Golden State" -> "state").
# Map those to a full "City Mascot" name so the canonical_event_id matches the
# nba_api-ingested events. ("LA Clippers" is already handled by matching.ALIASES.)
ARCHIVE_TEAM_FIX = {
    "Golden State": "Golden State Warriors",   # else -> "state"
    "Seventysixers": "Philadelphia 76ers",     # else -> "seventysixers"
    "Oklahoma City": "Oklahoma City Thunder",  # else -> "city"
    "NewJersey": "New Jersey Nets",            # -> "nets" (matches Brooklyn too)
}


# -- pure core (unit-testable, no I/O) --------------------------------------
def sbro_date_to_iso(mmdd, season_start_year: int) -> str | None:
    """``"1031"`` + 2023 -> ``"2023-10-31"``; ``"103"`` + 2023 -> ``"2024-01-03"``.

    SBRO stores the date as ``MMDD`` (no year) and a season straddles two years,
    so months in July..December belong to ``season_start_year`` and Jan..June to
    the next. Returns ``None`` for an unparseable cell.
    """
    s = str(mmdd).strip()
    if not s.isdigit() or len(s) < 3:
        return None
    month, day = int(s[:-2]), int(s[-2:])
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    year = season_start_year if month >= 7 else season_start_year + 1
    return f"{year:04d}-{month:02d}-{day:02d}"


def _team_name(label: str) -> str:
    """Map an SBRO label to a full team name, or pass it through if unmapped."""
    return SBRO_TEAMS.get((label or "").strip(), (label or "").strip())


def sbro_rows_to_records(rows: list[dict], season_start_year: int,
                         sport: str = "nba") -> list[dict]:
    """Pair visitor/home SBRO rows into ``ingest_odds`` feed records.

    Each game is a ``V`` row immediately followed by an ``H`` row sharing the
    same date. Rows are matched by walking the list and pairing every ``V`` with
    the next ``H``; any row missing its partner, a usable date, or a moneyline is
    skipped (``ingest_odds`` applies the vig guard after decimal conversion).
    ``ML`` is emitted verbatim so odds parsing stays single-sourced there.
    """
    records: list[dict] = []
    pending_visitor: dict | None = None
    unmapped: set[str] = set()

    for row in rows:
        vh = str(row.get("VH", "")).strip().upper()
        team = str(row.get("Team", "")).strip()
        if not team:
            continue
        if team not in SBRO_TEAMS:
            unmapped.add(team)
        if vh == "V":
            pending_visitor = row
            continue
        if vh != "H" or pending_visitor is None:
            pending_visitor = None        # neutral-site / malformed: reset pairing
            continue

        when = sbro_date_to_iso(row.get("Date"), season_start_year)
        v = pending_visitor
        pending_visitor = None
        if when is None:
            continue
        home_ml, away_ml = row.get("ML"), v.get("ML")
        if home_ml in (None, "", "NL", "nl") or away_ml in (None, "", "NL", "nl"):
            continue
        records.append({
            "sport": sport,
            "date": when,
            "home_team": _team_name(team),
            "away_team": _team_name(str(v.get("Team", "")).strip()),
            "home_close": home_ml,
            "away_close": away_ml,
        })

    if unmapped:
        log.warning("%d unmapped SBRO team label(s) passed through raw: %s",
                    len(unmapped), ", ".join(sorted(unmapped)))
    return records


def archive_date_to_iso(value) -> str | None:
    """``20111225`` / ``20111225.0`` -> ``"2011-12-25"``; ``None`` if unparseable."""
    try:
        s = str(int(float(value)))
    except (TypeError, ValueError):
        return None
    if len(s) != 8:
        return None
    return f"{s[:4]}-{s[4:6]}-{s[6:]}"


def archive_json_to_records(games: list[dict], sport: str = "nba") -> list[dict]:
    """Reshape the SBRO mirror's ``*_archive_10Y.json`` into feed records.

    Each element is already one pre-joined game with ``date`` (``YYYYMMDD``),
    ``home_team`` / ``away_team`` (mascot), and ``home_close_ml`` /
    ``away_close_ml`` (American). Junk rows (a ``0`` team or ``0`` moneyline,
    which the source uses for malformed/voided games) are dropped; the vig guard
    in ``ingest_odds`` is the final filter after decimal conversion.
    """
    out: list[dict] = []
    for g in games:
        when = archive_date_to_iso(g.get("date"))
        home = ARCHIVE_TEAM_FIX.get(str(g.get("home_team")).strip(), str(g.get("home_team")).strip())
        away = ARCHIVE_TEAM_FIX.get(str(g.get("away_team")).strip(), str(g.get("away_team")).strip())
        hc, ac = g.get("home_close_ml"), g.get("away_close_ml")
        if when is None or home in ("0", "", "None") or away in ("0", "", "None"):
            continue
        if hc in (None, "", 0) or ac in (None, "", 0):
            continue
        out.append({"sport": sport, "date": when, "home_team": home,
                    "away_team": away, "home_close": hc, "away_close": ac})
    return out


# -- I/O wrappers -----------------------------------------------------------
def read_rows(path: Path) -> list[dict]:
    """Read an SBRO export to a list of dict rows. CSV (stdlib) or .xlsx (openpyxl)."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # noqa: BLE001
            raise SystemExit("Reading .xlsx needs openpyxl (pip install openpyxl); "
                             "or export the sheet to CSV.") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(it)]
        return [dict(zip(header, row)) for row in it]
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert an SBRO NBA odds export -> ingest_odds feed JSON")
    parser.add_argument("input", help="SBRO export: classic .csv/.xlsx, or mirror archive .json")
    parser.add_argument("-o", "--output", help="feed JSON path (default: stdout)")
    parser.add_argument("--format", choices=("auto", "sbro", "archive"), default="auto",
                        help="'sbro' = classic two-row Excel/CSV; 'archive' = mirror "
                             "*_archive_10Y.json; 'auto' picks by extension")
    parser.add_argument("--season-start-year", type=int,
                        help="(sbro format only) year the season tipped off, e.g. 2023 for 2023-24")
    parser.add_argument("--sport", default="nba")
    args = parser.parse_args()

    path = Path(args.input)
    fmt = args.format
    if fmt == "auto":
        fmt = "archive" if path.suffix.lower() == ".json" else "sbro"

    if fmt == "archive":
        games = json.loads(path.read_text(encoding="utf-8"))
        records = archive_json_to_records(games, args.sport)
        log.info("Converted %d archive rows -> %d game records.", len(games), len(records))
    else:
        if args.season_start_year is None:
            parser.error("--season-start-year is required for the classic sbro format")
        rows = read_rows(path)
        records = sbro_rows_to_records(rows, args.season_start_year, args.sport)
        log.info("Converted %d SBRO rows -> %d game records.", len(rows), len(records))
    payload = json.dumps(records, indent=2)
    if args.output:
        Path(args.output).write_text(payload, encoding="utf-8")
        log.info("Wrote %s", args.output)
    else:
        print(payload)


if __name__ == "__main__":
    main()
